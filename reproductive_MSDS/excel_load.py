from openpyxl import load_workbook

# excel workbook 로드하기
def load_wb(file_name):
    wb = load_workbook(file_name)
    ws = wb.active
    return (wb, ws)

# 엑셀에서 카스번호 가져오기
def bring_casnum(ws, no):
    casnum = ws[f"D{no+1}"].value 
    return  casnum

# 엑셀에 원본 데이터 넣기
def insert_data(ws, data, no):
    ws[f"E{no+1}"].value = data       
