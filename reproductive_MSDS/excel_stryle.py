from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Color, Alignment

# 엑셀 로드
file_name = "reproductive_MSDS_원본추가.xlsx"
wb = load_workbook(file_name)
ws = wb.active

# 열 넓이 조절
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 15

# title 폰트, 정렬, 배경색
for cell in ws[1]:
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment("center", "center")
    cell.fill = PatternFill(fgColor=Color("FFFF00"), fill_type="solid")
    
    # E 열만 다른 색
    if cell.column == 5: # E열 문자대신 숫자로
        cell.fill = PatternFill(fgColor=Color("00FF00"), fill_type="solid")

# 틀고정
ws.freeze_panes = "E2"

# 저장
wb.save(file_name)
